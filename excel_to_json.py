import openpyxl
import json
import re
from pathlib import Path

EXCEL_FILE = "ПРАЙС ЛИСТ ЭНТЕХ от 31.08.23.xlsx"
OUT_JSON = "catalog.json"

def parse_sheet(ws, sheet_name):
    items = []
    print(f"Processing sheet: {sheet_name}")
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) < 3 or not row[0]:  # Проверяем, что есть хотя бы 3 столбца и первый не пустой
            continue
        
        model = str(row[0]).strip() if row[0] else None
        print(f"Row {row_num}: Model = '{model}'")  # DEBUG: смотрим, что находим
        
        if not model or 'NRG' not in model.upper():
            continue
        
        specs = str(row[2] or '') if len(row) > 2 else ''
        specs_lower = specs.lower()
        
        # Парсим характеристики
        power_match = re.search(r'(\d+)\s*вт', specs_lower)
        power = int(power_match.group(1)) if power_match else None
        
        lumens_match = re.search(r'(\d+)\s*лм', specs_lower)
        lumens = int(lumens_match.group(1)) if lumens_match else None
        
        ip_match = re.search(r'ip(\d{2})', specs_lower)
        ip = ip_match.group(1) if ip_match else None
        ip_rating = f"IP{ip}" if ip else None
        
        category = sheet_name.lower()
        
        # Создаём URL изображения (исправленный f-string)
        sanitized_model = re.sub(r'[\W]+', '-', model.lower())
        image_url = f"https://ene-rgy.ru/images/{sanitized_model}.jpg"
        
        item = {
            "model": model,
            "name": model,
            "power_w": power,
            "lumens": lumens,
            "ip_rating": ip_rating,
            "category": category,
            "image_url": image_url,
            "raw": specs[:200]
        }
        items.append(item)
        print(f"Added item: {model}")  # DEBUG
    
    print(f"Found {len(items)} items in sheet '{sheet_name}'")
    return items

def excel_to_json(excel_path, json_path):
    try:
        if not Path(excel_path).exists():
            print(f"ERROR: Excel file not found: {excel_path}")
            print("Available files:")
            for f in Path('.').glob('*'):
                print(f"  - {f}")
            return
        
        print(f"Loading Excel: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
        
        catalog = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_items = parse_sheet(ws, sheet_name)
            catalog.extend(sheet_items)
        
        unique_items = {item['model']: item for item in catalog}.values()
        
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(list(unique_items), f, ensure_ascii=False, indent=2)
        
        print(f"SUCCESS: Saved {len(unique_items)} unique items to {json_path}")
        
        if len(unique_items) == 0:
            print("WARNING: No items found. Check:")
            print("  1. Excel file structure (models should be in column A)")
            print("  2. Models should contain 'NRG' in name")
            print("  3. Try opening Excel manually to verify data")
            
    except Exception as e:
        print(f"ERROR processing Excel: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    excel_to_json(EXCEL_FILE, OUT_JSON)