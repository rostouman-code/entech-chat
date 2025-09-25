import openpyxl
import os
import json

# Загружаем Excel
wb = openpyxl.load_workbook("ПРАЙС ЛИСТ ЭНТЕХ от 31.08.23.xlsx")
ws = wb.active

# Папка для картинок
output_dir = "images"
os.makedirs(output_dir, exist_ok=True)

catalog = []

# Перебор строк начиная с 5-й
for row in range(5, ws.max_row + 1):
    model = ws[f"A{row}"].value
    desc = ws[f"C{row}"].value
    retail = ws[f"D{row}"].value
    wholesale = ws[f"E{row}"].value

    if not model:
        continue

    # Проверяем картинки (они могут быть в ячейке B)
    img_path = None
    for img in ws._images:
        # openpyxl хранит координаты как (row, col), начиная с 0
        if img.anchor._from.row == row - 1 and img.anchor._from.col == 1:  # колонка B
            img_path = os.path.join(output_dir, f"{model}.png")
            img.image.save(img_path)  # сохраняем файл
            break

    catalog.append({
        "model": model,
        "description": desc,
        "price_retail": retail,
        "price_wholesale": wholesale,
        "image": img_path or None
    })

# Сохраняем в JSON
with open("catalog.json", "w", encoding="utf-8") as f:
    json.dump(catalog, f, ensure_ascii=False, indent=2)

print("✅ catalog.json создан, фото сохранены в папку images/")
