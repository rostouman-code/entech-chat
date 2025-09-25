import openpyxl
import os
import json
import base64
from io import BytesIO
from PIL import Image as PILImage  # pip install pillow openpyxl

# Загружаем Excel
wb = openpyxl.load_workbook("ПРАЙС ЛИСТ ЭНТЕХ от 31.08.23.xlsx")
ws = wb.active

catalog = []

# Перебор строк начиная с 5-й
for row in range(5, ws.max_row + 1):
    model = ws[f"A{row}"].value
    desc = ws[f"C{row}"].value
    retail = ws[f"D{row}"].value
    wholesale = ws[f"E{row}"].value

    if not model:
        continue

    img_base64 = None
    for img in ws._images:
        if img.anchor._from.row == row - 1 and img.anchor._from.col == 1:  # колонка B
            try:
                img_bytes = img._data()
                pil_img = PILImage.open(BytesIO(img_bytes))
                buffered = BytesIO()
                pil_img.save(buffered, format="PNG")  # сохраняем в PNG
                img_base64 = "data:image/png;base64," + base64.b64encode(buffered.getvalue()).decode()
            except Exception as e:
                print(f"⚠️ Ошибка с картинкой {model}: {e}")
            break

    catalog.append({
        "model": model,
        "description": desc,
        "price_retail": retail,
        "price_wholesale": wholesale,
        "image_base64": img_base64
    })

# Сохраняем JSON
with open("catalog.json", "w", encoding="utf-8") as f:
    json.dump(catalog, f, ensure_ascii=False, indent=2)

print("✅ catalog.json создан, фотки встроены как base64")
